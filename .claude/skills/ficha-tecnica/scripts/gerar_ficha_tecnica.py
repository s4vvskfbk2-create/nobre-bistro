#!/usr/bin/env python3
"""Gerador de Ficha Técnica — padrão Nobre Bistrô / Sushi Boys.

Lê um JSON com a receita (ver reference/formato_dados.md) e produz um .xlsx com
3 abas: Ficha Técnica Padrão, Receita Operacional e Tabela Preços Dinâmica.
Todas as fórmulas são dinâmicas; a Custo Unit. da Aba 1 faz VLOOKUP na Aba 3.

Uso:
    python3 gerar_ficha_tecnica.py --dados receita.json --saida Ficha.xlsx
"""
import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl não instalado. Rode: pip install openpyxl")

# ---- Paleta / estilos (tema Nobre Bistrô) ----
VERDE_ESCURO = "2D5016"
VERDE_CLARO = "90EE90"
AMARELO = "FFFFCC"

FILL_TITULO = PatternFill("solid", fgColor=VERDE_ESCURO)
FILL_HEADER = PatternFill("solid", fgColor=VERDE_CLARO)
FILL_FORMULA = PatternFill("solid", fgColor=AMARELO)

FONT_TITULO = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_LABEL = Font(bold=True)
FONT_BASE = Font(name="Calibri", size=11)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_PRECOS = "Tabela Preços Dinâmica"


def _title(ws, last_col, texto):
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = texto
    c.fill = FILL_TITULO
    c.font = FONT_TITULO
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _header_cell(cell, texto):
    cell.value = texto
    cell.fill = FILL_HEADER
    cell.font = FONT_HEADER
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER


def aba_ficha_tecnica(wb, d):
    ws = wb.active
    ws.title = "Ficha Técnica Padrão"
    _title(ws, "G", f"FICHA TÉCNICA - {d.get('prato', '')}")

    ws["A2"] = "Item:"; ws["A2"].font = FONT_LABEL
    ws["C2"] = d.get("prato", "")
    ws["E2"] = "Classificação:"; ws["E2"].font = FONT_LABEL
    ws["G2"] = d.get("classificacao", "")
    ws["A3"] = "Setor:"; ws["A3"].font = FONT_LABEL
    ws["C3"] = d.get("setor", "Alimentos e Bebidas")
    ws["E3"] = "Tamanho da Receita:"; ws["E3"].font = FONT_LABEL
    ws["G3"] = d.get("porcoes", 1)
    ws["A4"] = "Tamanho da Porção (g):"; ws["A4"].font = FONT_LABEL
    ws["C4"] = d.get("tamanho_porcao_g", 0)

    # Cabeçalho da tabela de ingredientes (linha 6)
    headers = ["Ingrediente", "Qtd. Líquida", "Und.", "Aprov.",
               "Qtd. Bruta", "Custo Unit.", "Custo Total"]
    for i, h in enumerate(headers):
        _header_cell(ws.cell(row=6, column=1 + i), h)

    ingredientes = d.get("ingredientes", [])
    row = 7
    for ing in ingredientes:
        nome = ing.get("nome", "")
        ws.cell(row=row, column=1, value=nome).border = BORDER
        # Qtd. Líquida = Qtd. Bruta * Aprov.
        ws.cell(row=row, column=2, value=f"=E{row}*D{row}").border = BORDER
        ws.cell(row=row, column=3, value=ing.get("unidade", "")).border = BORDER
        ws.cell(row=row, column=4, value=ing.get("aprov", 1)).border = BORDER
        ws.cell(row=row, column=5, value=ing.get("qtd_bruta", 0)).border = BORDER
        # Custo Unit. via VLOOKUP na Tabela de Preços (dinâmico)
        cu = ws.cell(
            row=row, column=6,
            value=(f"=IFERROR(VLOOKUP(A{row},'{SHEET_PRECOS}'!$A:$C,3,FALSE),0)"),
        )
        cu.border = BORDER
        cu.number_format = "#,##0.00"
        # Custo Total = Custo Unit. * Qtd. Bruta
        ct = ws.cell(row=row, column=7, value=f"=F{row}*E{row}")
        ct.border = BORDER
        ct.number_format = "#,##0.00"
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAL CUSTO").font = FONT_LABEL
    total = ws.cell(row=total_row, column=7,
                    value=f"=SUM(G7:G{row - 1})")
    total.fill = FILL_FORMULA
    total.font = FONT_LABEL
    total.number_format = "#,##0.00"
    total_ref = f"G{total_row}"

    # Bloco de cálculos
    base = total_row + 2
    ws.cell(row=base, column=1, value="CÁLCULOS").font = FONT_TITULO
    ws.cell(row=base, column=1).fill = FILL_TITULO

    def calc(r, label, formula, fmt="#,##0.00"):
        ws.cell(row=r, column=1, value=label).font = FONT_LABEL
        cell = ws.cell(row=r, column=3, value=formula)
        cell.fill = FILL_FORMULA
        cell.number_format = fmt

    pv = base + 1
    ws.cell(row=pv, column=1, value="Preço de Venda (defina/pesquise):").font = FONT_LABEL
    pvc = ws.cell(row=pv, column=3, value=d.get("preco_venda", 0))
    pvc.fill = FILL_FORMULA
    pvc.number_format = "#,##0.00"
    pv_ref = f"C{pv}"

    calc(pv + 1, "CMV %", f"=IFERROR(({total_ref}/{pv_ref})*100,0)", '0.00"%"')
    calc(pv + 2, "Margem Bruta R$", f"={pv_ref}-{total_ref}")
    calc(pv + 3, "Margem Bruta %",
         f"=IFERROR((C{pv + 2}/{pv_ref})*100,0)", '0.00"%"')
    calc(pv + 4, "Markup", f"=IFERROR({pv_ref}/{total_ref},0)", '0.00"x"')
    calc(pv + 5, "Custo por Porção",
         f"=IFERROR({total_ref}/G3,0)")

    # Psicologia de preço
    psi = pv + 7
    ws.cell(row=psi, column=1, value="PSICOLOGIA DE PREÇO").font = FONT_TITULO
    ws.cell(row=psi, column=1).fill = FILL_TITULO
    ws.cell(row=psi + 1, column=1, value="Preço Esperado (mercado):").font = FONT_LABEL
    pe = ws.cell(row=psi + 1, column=3, value=d.get("preco_esperado_mercado", 0))
    pe.fill = FILL_FORMULA
    pe.number_format = "#,##0.00"
    ws.cell(row=psi + 2, column=1, value="Preço Psicológico (,90):").font = FONT_LABEL
    pp = ws.cell(row=psi + 2, column=3, value=f"=INT({pv_ref})-0.10")
    pp.fill = FILL_FORMULA
    pp.number_format = "#,##0.00"
    ws.cell(row=psi + 3, column=1, value="Preço Final Recomendado:").font = FONT_LABEL
    pf = ws.cell(row=psi + 3, column=3, value=f"=C{psi + 2}")
    pf.fill = FILL_FORMULA
    pf.number_format = "#,##0.00"

    # Rodapé
    foot = psi + 5
    ws.cell(row=foot, column=1,
            value=f"Elaborado por: {d.get('elaborado_por', '')} | "
                  f"Próxima revisão: ___/___/______").font = Font(italic=True, size=9)

    widths = {"A": 28, "B": 13, "C": 13, "D": 9, "E": 12, "F": 12, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def aba_operacional(wb, d):
    ws = wb.create_sheet("Receita Operacional")
    _title(ws, "C", f"RECEITA OPERACIONAL - {d.get('prato', '')}")

    ws["A2"] = "Prato:"; ws["A2"].font = FONT_LABEL
    ws["C2"] = d.get("prato", "")
    ws["A3"] = "Rendimento:"; ws["A3"].font = FONT_LABEL
    ws["C3"] = f"{d.get('porcoes', '')} porções"
    ws["A4"] = "Tempo de Preparo:"; ws["A4"].font = FONT_LABEL
    ws["C4"] = f"{d.get('tempo_preparo_min', '')} minutos"

    ws["A6"] = "INGREDIENTES (UNIDADES DOMÉSTICAS)"
    ws["A6"].font = FONT_LABEL
    for i, h in enumerate(["Ingrediente", "Quantidade", "Unidade"]):
        _header_cell(ws.cell(row=7, column=1 + i), h)

    row = 8
    for ing in d.get("ingredientes_operacional", []):
        ws.cell(row=row, column=1, value=ing.get("nome", "")).border = BORDER
        ws.cell(row=row, column=2, value=ing.get("quantidade", "")).border = BORDER
        ws.cell(row=row, column=3, value=ing.get("unidade", "")).border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="MODO DE PREPARO").font = FONT_LABEL
    row += 1
    for i, passo in enumerate(d.get("modo_preparo", []), start=1):
        ws.cell(row=row, column=1, value=f"{i}. {passo}")
        row += 1

    dicas = d.get("dicas", [])
    if dicas:
        row += 1
        ws.cell(row=row, column=1, value="DICAS").font = FONT_LABEL
        row += 1
        for dica in dicas:
            ws.cell(row=row, column=1, value=f"- {dica}")
            row += 1

    val = d.get("validade")
    if val:
        row += 1
        ws.cell(row=row, column=1, value="VALIDADE").font = FONT_LABEL
        row += 1
        if val.get("geladeira_dias") is not None:
            ws.cell(row=row, column=1,
                    value=f"Geladeira (0-4°C): {val['geladeira_dias']} dias")
            row += 1
        if val.get("freezer_dias") is not None:
            ws.cell(row=row, column=1,
                    value=f"Freezer (-18°C): {val['freezer_dias']} dias")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18


def aba_precos(wb, d):
    ws = wb.create_sheet(SHEET_PRECOS)
    _title(ws, "F", "TABELA DE PREÇOS - GESTÃO DE CUSTOS")
    ws["A2"] = "Atualizado em: ___/___/______"
    ws["A2"].font = Font(italic=True, size=9)

    headers = ["Ingrediente", "Unidade", "Preço Unit",
               "Data Compra", "Fornecedor", "Observações"]
    for i, h in enumerate(headers):
        _header_cell(ws.cell(row=4, column=1 + i), h)

    row = 5
    for ing in d.get("ingredientes", []):
        ws.cell(row=row, column=1, value=ing.get("nome", "")).border = BORDER
        ws.cell(row=row, column=2, value=ing.get("unidade", "")).border = BORDER
        pc = ws.cell(row=row, column=3, value=ing.get("custo_unit", 0))
        pc.fill = FILL_FORMULA
        pc.number_format = "#,##0.00"
        pc.border = BORDER
        ws.cell(row=row, column=4, value=ing.get("data_compra", "")).border = BORDER
        ws.cell(row=row, column=5, value=ing.get("fornecedor", "")).border = BORDER
        ws.cell(row=row, column=6, value=ing.get("obs", "")).border = BORDER
        row += 1

    widths = {"A": 28, "B": 10, "C": 12, "D": 14, "E": 16, "F": 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    ap = argparse.ArgumentParser(description="Gera ficha técnica .xlsx")
    ap.add_argument("--dados", required=True, help="JSON da receita")
    ap.add_argument("--saida", required=True, help="Arquivo .xlsx de saída")
    args = ap.parse_args()

    with open(args.dados, encoding="utf-8") as f:
        d = json.load(f)

    wb = Workbook()
    aba_ficha_tecnica(wb, d)
    aba_operacional(wb, d)
    aba_precos(wb, d)
    wb.save(args.saida)
    print(f"OK: {args.saida} gerado com 3 abas.")


if __name__ == "__main__":
    main()
